from bs4 import BeautifulSoup
import requests
import io
from pdfminer.converter import TextConverter
from pdfminer.pdfinterp import PDFPageInterpreter, PDFResourceManager
from pdfminer.pdfpage import PDFPage
from pdfminer.layout import LAParams
from openpyxl import load_workbook
import datetime

# Get HTML code of the page for applicants
html = requests.get("https://www.vhutein.ru/applicants/documents/")
# Initialize BeautifulSoup with obtained HTML
soup = BeautifulSoup(html.text, "lxml")
# Find in obtained HTML first "a" element with class "link_doc"
link = soup.find("a", {"class": "link_doc"})
# Parse and format date publication of the file
full_date = link.text[len(link.text) - 16:len(link.text) - 8]
# Format date for new file names
short_date = full_date[0:5].replace(".", "")
# Format date for Excel date operations
current_date = datetime.datetime(int("20" + full_date[6:8]), int(full_date[3:5]), int(full_date[0:2]))

# Get the PDF file with ranked list of applicants
file = requests.get("https://www.vhutein.ru" + link.attrs["href"])
# Initialize binary io stream
io_binary = io.BytesIO()
# Write obtained PDF in binary format to the binary io stream
io_binary.write(file.content)

# Create or open file to save source PDF
pdf = open("saved_files\\file_" + short_date + ".pdf", "wb")
# Write date
pdf.write(file.content)
# Close file
pdf.close()

# Initialize PDFResourceManager
resource_manager = PDFResourceManager()
# Initialize string io stream
io_stream = io.StringIO()
# Initialize TextConverter with special parameters to read data in correct order
converter = TextConverter(resource_manager, io_stream, codec="utf-8", laparams=LAParams(boxes_flow=None,
                                                                                        char_margin=1.0,
                                                                                        line_margin=0.1))
# Initialize PDFPageInterpreter
page_interpreter = PDFPageInterpreter(resource_manager, converter)

# Read and parse PDF pages and save it in the string io stream
for page in PDFPage.get_pages(io_binary,
                              caching=True,
                              check_extractable=True):
    # Parse PDF page
    page_interpreter.process_page(page)

# Close TextConverter
converter.close()

# Close binary io stream
io_binary.close()

# Get and format the parse result
text = io_stream.getvalue().replace("\n\n", "\n").replace("\f", "")
# Clear source PDF data
io_stream.truncate(0)
# Set cursor point to 0
io_stream.seek(0)
# Write formatted PDF data
io_stream.write(text)
# Set cursor point to 0
io_stream.seek(0)

# Create or open file to save parsed and formatted PDF in TXT format
txt = open("saved_files\\file_" + short_date + ".txt", "w")
# Write data
txt.write(text)
# Close file
txt.close()


# Function for format data to correct type
def reformat_data(data, data_type):
    data = data.replace("\n", "")
    if data_type == "creative" or data_type == "russian" or data_type == "literature":
        if data == "-" or data == "–":
            return 0
        return int(data)
    return data


# Function for get mean with ignore zero values
def get_mean(arr):
    new_arr = [elem for elem in arr if elem != 0]
    if len(new_arr) == 0:
        return 0
    return sum(new_arr) / len(new_arr)


fact = []
creatives = []
russians = []
literatures = []
applicants = []

while True:
    line = io_stream.readline()
    if not line:
        break
    if line.find("/ФД/") != -1:
        direction = reformat_data(io_stream.readline(), "direction")
        creative = reformat_data(io_stream.readline(), "creative")
        russian = reformat_data(io_stream.readline(), "russian")
        literature = reformat_data(io_stream.readline(), "literature")
        io_stream.readline()
        status = reformat_data(io_stream.readline(), "status")
        if status != "да" and status != "–":
            continue

        if creative != 0:
            creatives.append(creative)
        if russian != 0:
            russians.append(russian)
        if literature != 0:
            literatures.append(literature)

        applicants.append([direction, creative, russian, literature])

        fact.append([direction, get_mean([creative, russian, literature])])

    if line.find("/ПИ/") != -1:
        break

# Close string io stream
io_stream.close()

fact.sort(key=lambda elem: elem[1], reverse=True)
creatives.sort(reverse=True)
mean_creative = get_mean(creatives)
russians.sort(reverse=True)
mean_russian = get_mean(russians)
literatures.sort(reverse=True)
mean_literature = get_mean(literatures)
mean = []

for applicant in applicants:
    mean.append([applicant[0], get_mean([applicant[1] if applicant[1] != 0 else mean_creative,
                                         applicant[2] if applicant[2] != 0 else mean_russian,
                                         applicant[3] if applicant[3] != 0 else mean_literature])])

mean.sort(key=lambda elem: elem[1], reverse=True)

full = []

for applicant in applicants:
    if applicant[1] != 0 and applicant[2] != 0 and applicant[3] != 0:
        full.append([applicant[0], get_mean([applicant[1], applicant[2], applicant[3]])])

full.sort(key=lambda elem: elem[1], reverse=True)


def select_gd(arr):
    new_arr = []
    num_dm = 50
    num_gd = 100
    num_del_dm = 0
    for elem in arr:
        if num_del_dm != num_dm:
            if elem[0] == "ДМ" or elem[0] == "1 ДМ / 2 ГД":
                num_del_dm += 1
            else:
                new_arr.append(elem[1])
        else:
            if elem[0] != "ДМ":
                new_arr.append(elem[1])
    if len(new_arr) > num_gd:
        new_arr = new_arr[0:num_gd]
    return new_arr


gd_fact = select_gd(fact)
gd_mean = select_gd(mean)
gd_full = select_gd(full)

rank = 0
avg_value = 68
if gd_full[len(gd_full) - 1] <= avg_value:
    for val in range(len(gd_full)):
        if gd_full[val] <= avg_value:
            rank = val + 1
            break


def find_row(sheet, date):
    current_row = sheet.max_row + 1
    for row in range(1, current_row):
        if sheet[row][0].value == date:
            current_row = row
            break
    return current_row


def write_statistics(sheet, date, arr):
    current_row = find_row(sheet, date)
    sheet[current_row][0].value = date
    sheet[current_row][0].number_format = "mm-dd-yy"
    sheet[current_row][1].value = arr[0]
    sheet[current_row][1].number_format = "0.00"
    sheet[current_row][2].value = get_mean(arr)
    sheet[current_row][2].number_format = "0.00"
    sheet[current_row][3].value = arr[len(arr) - 1]
    sheet[current_row][3].number_format = "0.00"


def write_rank(sheet, date, value):
    current_row = find_row(sheet, date)
    sheet[current_row][0].value = date
    sheet[current_row][0].number_format = "mm-dd-yy"
    sheet[current_row][1].value = value


# Open Excel file
workbook = load_workbook(filename='statistics.xlsx')
# Write data for all arrays in Excel file
write_statistics(workbook["fact"], current_date, gd_fact)
write_statistics(workbook["mean"], current_date, gd_mean)
write_statistics(workbook["full"], current_date, gd_full)
write_statistics(workbook["test"], current_date, creatives)
write_rank(workbook["rank"], current_date, rank)
# Save Excel file
workbook.save(filename='statistics.xlsx')
