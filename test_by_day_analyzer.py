import datetime
from openpyxl import load_workbook


def reformat_data(data):
    data = data.replace("\n", "")
    if data == "-" or data == "–":
        return 0
    return int(data)


test_by_day = []

for month in range(7, 9):
    for day in range(1, 32):
        date = ("0" + str(day))[-2:] + ("0" + str(month))[-2:]
        try:
            file = open("saved_files\\file_" + date + ".txt", "r")
            test = []
            while True:
                line = file.readline()
                if not line:
                    break
                if line.find("/ФД/") != -1:
                    file.readline()
                    test_data = reformat_data(file.readline())
                    if test_data != 0:
                        test.append(test_data)
                if line.find("/ПИ/") != -1:
                    break
            test_by_day.append([date, test])
            file.close()
        except FileNotFoundError:
            continue

dates_count = len(test_by_day)
for i in range(1, dates_count):
    j = dates_count - i
    for value in test_by_day[j-1][1]:
        test_by_day[j][1].remove(value)

tbyd = []
tbyw = [["Понедельник", []],
        ["Вторник", []],
        ["Среда", []],
        ["Четверг", []],
        ["Пятница", []],
        ["Суббота", []],
        ["Воскресенье", []]]
first_day = datetime.datetime(2021, int(test_by_day[0][0][2:4]), int(test_by_day[0][0][0:2]))
for test_day in test_by_day:
    if len(test_day[1]) != 0:
        some_date = datetime.datetime(2021, int(test_day[0][2:4]), int(test_day[0][0:2]))
        some_value = sum(test_day[1]) / len(test_day[1])
        tbyd.append([some_date, some_value])
        if some_date != first_day:
            tbyw[some_date.weekday()][1].append(some_value)


def find_row(sheet, current_date):
    current_row = sheet.max_row + 1
    for row in range(1, current_row):
        if sheet[row][0].value == current_date:
            current_row = row
            break
    return current_row


def write_tbyd(sheet, current_date, current_value):
    current_row = find_row(sheet, current_date)
    sheet[current_row][0].value = current_date
    sheet[current_row][0].number_format = "mm-dd-yy"
    sheet[current_row][1].value = current_value
    sheet[current_row][1].number_format = "0.00"


def write_tbyw(sheet, current_date, current_value):
    current_row = find_row(sheet, current_date)
    sheet[current_row][0].value = current_date
    if len(current_value) != 0:
        sheet[current_row][1].value = sum(current_value) / len(current_value)
    else:
        sheet[current_row][1].value = 0
    sheet[current_row][1].number_format = "0.00"


workbook = load_workbook(filename='statistics.xlsx')
for day in tbyd:
    write_tbyd(workbook["tbyd"], day[0], day[1])
for day in tbyw:
    write_tbyw(workbook["tbyw"], day[0], day[1])
workbook.save(filename='statistics.xlsx')

